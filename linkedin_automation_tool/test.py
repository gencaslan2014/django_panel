from openpyxl import load_workbook
import django
import os
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'linkedin_automation_tool.settings')
django.setup()
from tool.models import *
wb = load_workbook(r'C:\Users\Administrator\PycharmProjects\django_panel\linkedin_automation_tool\Outreaching MS Patners (1).xlsx')
ws = wb['SalesNavigator']
for j,row in enumerate(ws.iter_rows(min_row=2)):
    print(f"Saving Record={j}")
    try:
        print(row[1].value)
        print(row[0].value)
        leads = Leads()
        leads.name = row[1].value+" "+row[2].value
        leads.last_name = row[2].value
        leads.first_name = row[1].value
        leads.title = row[3].value
        leads.email = row[4].value
        leads.phone_number = row[5].value
        leads.linkedin = row[6].value
        leads.employees = row[7].value
        leads.company_website = row[8].value
        leads.company_name = row[9].value
        leads.location = row[10].value
        leads.linkedin_sales_navigator = row[11].value
        leads.save()
        # break
    except Exception as e:
        print(e)

